import streamlit as st
import pdfplumber
import re
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

st.set_page_config(page_title="Extractor MO H&P AUDATEX", page_icon="🔧", layout="wide")

st.markdown("""
<style>
.main-header{background:linear-gradient(135deg,#1a1a2e 0%,#16213e 50%,#0f3460 100%);
    padding:2rem;border-radius:12px;margin-bottom:2rem;text-align:center;}
.main-header h1{color:#e94560;margin:0;font-size:2rem;}
.main-header p{color:#a8b2d8;margin:.4rem 0 0;font-size:.95rem;}
.metric-box{background:white;border-radius:8px;padding:.8rem 1rem;
    text-align:center;box-shadow:0 2px 6px rgba(0,0,0,.08);}
.metric-value{font-size:1.6rem;font-weight:800;color:#0f3460;}
.metric-label{font-size:.78rem;color:#6c757d;margin-top:.2rem;}
.sec-title{font-size:1.05rem;font-weight:700;color:#1a1a2e;margin-bottom:.5rem;}
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="main-header">
  <h1>🔧 Extractor AUDATEX – Carrocería</h1>
  <p>Extrae partidas de <b>Mano de Obra Hojal/Mecánica</b> y <b>Pintura de Carrocería</b> · Exporta a Excel por N° de Orden</p>
</div>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# EXTRACCIÓN
# ══════════════════════════════════════════════════════════════════════════════

def parse_price(s: str) -> float:
    return float(s.replace(",", ""))

RE_ITEM    = re.compile(r'^(\S+)\s+(.+?)\s+[\d.]+\s*(?:[R*]+)?\s*\$([\d,]+\.?\d*)\*?$')
RE_TIEMPO  = re.compile(r'^(TIEMPO\s+DE\s+PREP\..+?)\s+[\d.]+\s*R\s+\$([\d,]+\.?\d*)', re.I)
RE_MO_A    = re.compile(r'^Mano\s+de\s+Obra', re.I)
RE_MO_B    = re.compile(r'^Hojal/Mec', re.I)
RE_MO_TBL  = re.compile(r'NR\s+Operaci.+Trabajo\s+UT\s+Precio', re.I)
RE_MO_STOP = re.compile(r'^Total\s+(Unidades|M\.O\.)', re.I)
RE_PIN_HDR = re.compile(r'^PINTURA\s+DE\s+CARROCER', re.I)
RE_PIN_TBL = re.compile(r'NR\s+Operaci.+UT\s+Precio', re.I)
RE_PIN_STP = re.compile(r'^(RESUMEN\s+M\.O|Total\s+de\s+Horas)', re.I)
SKIP       = {'TOTAL','SUBTOTAL','SUMA','IVA','RESUMEN','PIEZAS'}


def extract_audatex(pdf_bytes: bytes) -> dict:
    mo_items, pin_items, meta = [], [], {}

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        lines = []
        for page in pdf.pages:
            lines.extend((page.extract_text() or "").splitlines())

    # ── metadata + resumen pintura ─────────────────────────────────────────────
    for line in lines:
        s = line.strip()
        def g(pat): m=re.search(pat,line); return m.group(1).strip() if m else None

        if "Referencia Interna:" in line and "num_orden" not in meta:
            meta["num_orden"]   = g(r'Referencia Interna:\s*(\S+)') or "–"
        if "Número de Expediente:" in line and "expediente" not in meta:
            meta["expediente"]  = g(r'Número de Expediente:\s*(\S+)') or "–"
        if "Número de Reclamo:" in line and "reclamo" not in meta:
            meta["reclamo"]     = g(r'Número de Reclamo:\s*(\S+)') or "–"
        if "Taller de Reparación:" in line and "taller" not in meta:
            meta["taller"]      = g(r'Taller de Reparación:\s*(.+)') or "–"
        if "Fabricante:" in line and "fabricante" not in meta:
            meta["fabricante"]  = g(r'Fabricante:\s*(.+)') or ""
        if re.search(r'^Modelo:\s*\S', line) and "modelo" not in meta:
            meta["modelo"]      = g(r'Modelo:\s*(.+)') or ""
        if "No. VIN Visual:" in line and "vin" not in meta:
            meta["vin"]         = g(r'No\. VIN Visual:\s*(\S+)') or "–"

        # Total MO Hojal
        if "Total M.O. Hojal" in line:
            m2 = re.search(r'\$([\d,]+\.?\d*)', line)
            if m2: meta["total_mo"] = parse_price(m2.group(1))

        # Resumen pintura: TIEMPO M.O  100.3$2,607.80
        if re.match(r'^TIEMPO\s+M\.O\b', s):
            m2 = re.search(r'\$([\d,]+\.?\d*)', s)
            if m2: meta["tiempo_mo_pintura"] = parse_price(m2.group(1))

        # TIEMPO PREPARACION  23.6 $613.60
        if re.match(r'^TIEMPO\s+PREPARACION', s):
            m2 = re.search(r'\$([\d,]+\.?\d*)', s)
            if m2: meta["tiempo_prep_pintura"] = parse_price(m2.group(1))

        # TOTAL M.O. PINTURA  123.90$3,221.40
        if re.match(r'^TOTAL\s+M\.O\.\s+PINTURA', s):
            m2 = re.search(r'\$([\d,]+\.?\d*)', s)
            if m2: meta["total_mo_pintura"] = parse_price(m2.group(1))

        # MATERIALES POR SUPERFICIE $5,478.38
        if re.match(r'^MATERIALES\s+POR\s+SUPERFICIE', s):
            m2 = re.search(r'\$([\d,]+\.?\d*)', s)
            if m2: meta["mat_por_superficie"] = parse_price(m2.group(1))

        # CONSTANTE MATERIAL $335.99
        if re.match(r'^CONSTANTE\s+MATERIAL', s):
            m2 = re.search(r'\$([\d,]+\.?\d*)', s)
            if m2: meta["constante_material"] = parse_price(m2.group(1))

        # TOTAL MATERIALES $5,814.37
        if re.match(r'^TOTAL\s+MATERIALES', s):
            m2 = re.search(r'\$([\d,]+\.?\d*)', s)
            if m2: meta["total_materiales"] = parse_price(m2.group(1))

    # ── state machine ──────────────────────────────────────────────────────────
    NONE, MO_PREHDR, MO_WAIT, MO, PIN_WAIT, PIN = range(6)
    state = NONE

    for line in lines:
        s = line.strip()
        if not s:
            continue

        if RE_PIN_HDR.match(s):
            state = PIN_WAIT; continue
        if state == PIN_WAIT:
            if RE_PIN_TBL.search(s): state = PIN
            continue
        if state == PIN:
            if RE_PIN_STP.match(s): state = NONE; continue
            if s.startswith("-"): continue
            m = RE_ITEM.match(s)
            if m and m.group(1).upper() not in SKIP:
                pin_items.append({"nr": m.group(1), "descripcion": m.group(2).strip(),
                                  "precio": parse_price(m.group(3))})
            continue

        if RE_MO_A.match(s): state = MO_PREHDR; continue
        if state == MO_PREHDR:
            if RE_MO_B.match(s): state = MO_WAIT; continue
            if RE_MO_TBL.search(s): state = MO; continue
            continue
        if state == MO_WAIT:
            if RE_MO_TBL.search(s): state = MO
            continue
        if state == MO:
            if RE_MO_STOP.match(s): state = NONE; continue
            if RE_PIN_HDR.match(s): state = PIN_WAIT; continue
            mt = RE_TIEMPO.match(s)
            if mt:
                mo_items.append({"nr": "", "descripcion": mt.group(1).strip(),
                                 "precio": parse_price(mt.group(2))}); continue
            m = RE_ITEM.match(s)
            if m and m.group(1).upper() not in SKIP:
                mo_items.append({"nr": m.group(1), "descripcion": m.group(2).strip(),
                                 "precio": parse_price(m.group(3))})

    return {"mo": mo_items, "pintura": pin_items, "meta": meta}


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def build_excel(all_data: list, filenames: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Partidas"

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 52
    ws.column_dimensions["E"].width = 16

    THIN   = Side(style="thin",   color="CCCCCC")
    THICK  = Side(style="medium", color="888888")
    BDR    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    BDR_TOP= Border(left=THIN, right=THIN, top=THICK, bottom=THIN)
    CTR    = Alignment(horizontal="center", vertical="center")
    LFT    = Alignment(horizontal="left",   vertical="center")
    RGT    = Alignment(horizontal="right",  vertical="center")
    MONEY  = '#,##0.00'

    # Fills
    HDR_FILL   = PatternFill("solid", fgColor="0F3460")
    MO_FILL    = PatternFill("solid", fgColor="D6E4F7")
    MO_ALT     = PatternFill("solid", fgColor="EBF3FB")
    PIN_FILL   = PatternFill("solid", fgColor="FADADD")
    PIN_ALT    = PatternFill("solid", fgColor="FDEEF0")
    GRP_FILL   = PatternFill("solid", fgColor="1A1A2E")
    # Summary row fills
    SUB1_FILL  = PatternFill("solid", fgColor="16213E")   # Tiempo MO / prep / total MO pintura
    SUB2_FILL  = PatternFill("solid", fgColor="1B3A2D")   # Materiales
    SUB3_FILL  = PatternFill("solid", fgColor="0B2A40")   # Total MO + Materiales
    GRAND_FILL = PatternFill("solid", fgColor="0B0B1A")

    W_FONT    = Font(color="FFFFFF", bold=True, size=10)
    W_FONT_SM = Font(color="FFFFFF", bold=True, size=9)
    W_FONT_IT = Font(color="DDDDDD", italic=True, size=9)
    BOLD      = Font(bold=True, size=10)

    # Column headers
    headers = ["N° ORDEN", "SECCIÓN", "NR / POS.", "TRABAJO / DESCRIPCIÓN", "PRECIO ($)"]
    ws.row_dimensions[1].height = 22
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.font=W_FONT; c.fill=HDR_FILL; c.alignment=CTR; c.border=BDR
    ws.freeze_panes = "A2"

    row = 2
    grand_mo_total  = 0.0
    grand_pin_total = 0.0

    def summary_row(label, value, fill, font, is_first=False):
        nonlocal row
        ws.row_dimensions[row].height = 16
        bdr = BDR_TOP if is_first else BDR
        ws.merge_cells(f"A{row}:D{row}")
        ct = ws.cell(row, 1, f"  {label}")
        cv = ws.cell(row, 5, value)
        ct.font=font; ct.fill=fill; ct.alignment=LFT; ct.border=bdr
        cv.font=font; cv.fill=fill; cv.alignment=RGT; cv.border=bdr
        cv.number_format=MONEY
        # fill merged cells 2-4
        for col in range(2, 5):
            c = ws.cell(row, col)
            c.fill=fill; c.border=bdr
        row += 1

    for fname, data in zip(filenames, all_data):
        meta  = data["meta"]
        mo    = data["mo"]
        pin   = data["pintura"]
        n_ord = meta.get("num_orden", "–")
        veh   = f"{meta.get('fabricante','')} {meta.get('modelo','')}".strip()

        # ── Group header ──────────────────────────────────────────────────────
        ws.merge_cells(f"A{row}:E{row}")
        ws.row_dimensions[row].height = 18
        c = ws.cell(row, 1,
            f"  N° ORDEN: {n_ord}   |   Exp: {meta.get('expediente','–')}   |   "
            f"{veh}   |   VIN: {meta.get('vin','–')}   |   Taller: {meta.get('taller','–')}")
        c.font=W_FONT; c.fill=GRP_FILL; c.alignment=LFT
        c.border=Border(left=THICK,right=THICK,top=THICK,bottom=THICK)
        row += 1

        # ── MO rows ───────────────────────────────────────────────────────────
        mo_sum = 0.0
        for i, item in enumerate(mo):
            ws.row_dimensions[row].height = 16
            fill = MO_FILL if i%2==0 else MO_ALT
            ws.cell(row,1,n_ord).alignment=CTR
            ws.cell(row,2,"Mano de Obra Hojal/Mecánica").alignment=LFT
            ws.cell(row,3,item["nr"]).alignment=CTR
            ws.cell(row,4,item["descripcion"]).alignment=LFT
            p=ws.cell(row,5,item["precio"]); p.number_format=MONEY; p.alignment=RGT
            for col in range(1,6):
                c=ws.cell(row,col); c.fill=fill; c.border=BDR_TOP if i==0 else BDR
                if col==1: c.font=Font(bold=True,size=9,color="0F3460")
                elif col==2: c.font=Font(size=9,color="0F3460")
                else: c.font=Font(size=9)
            mo_sum += item["precio"]; row += 1

        # MO subtotal (from PDF)
        if mo:
            total_mo = meta.get("total_mo", mo_sum)
            summary_row(f"SUBTOTAL  Mano de Obra Hojal/Mecánica  ·  N° Orden {n_ord}",
                        total_mo, SUB1_FILL, W_FONT_SM, is_first=True)
            grand_mo_total += total_mo

        # ── Pintura rows ──────────────────────────────────────────────────────
        pin_sum = 0.0
        for i, item in enumerate(pin):
            ws.row_dimensions[row].height = 16
            fill = PIN_FILL if i%2==0 else PIN_ALT
            ws.cell(row,1,n_ord).alignment=CTR
            ws.cell(row,2,"Pintura de Carrocería").alignment=LFT
            ws.cell(row,3,item["nr"]).alignment=CTR
            ws.cell(row,4,item["descripcion"]).alignment=LFT
            p=ws.cell(row,5,item["precio"]); p.number_format=MONEY; p.alignment=RGT
            for col in range(1,6):
                c=ws.cell(row,col); c.fill=fill; c.border=BDR_TOP if i==0 else BDR
                if col==1: c.font=Font(bold=True,size=9,color="8B0000")
                elif col==2: c.font=Font(size=9,color="8B0000")
                else: c.font=Font(size=9)
            pin_sum += item["precio"]; row += 1

        # ── Pintura summary block ─────────────────────────────────────────────
        if pin:
            tiempo_mo   = meta.get("tiempo_mo_pintura",  pin_sum)
            tiempo_prep = meta.get("tiempo_prep_pintura", 0.0)
            total_mo_p  = meta.get("total_mo_pintura",   tiempo_mo + tiempo_prep)
            mat_sup     = meta.get("mat_por_superficie",  0.0)
            constante   = meta.get("constante_material",  0.0)
            total_mat   = meta.get("total_materiales",    mat_sup + constante)
            total_mo_mat = total_mo_p + total_mat

            # Tiempo M.O.
            summary_row("Tiempo M.O.",        tiempo_mo,   SUB1_FILL, W_FONT_SM, is_first=True)
            # Tiempo Preparación
            summary_row("Tiempo Preparación", tiempo_prep, SUB1_FILL, W_FONT_IT)
            # Total M.O. Pintura
            summary_row("Total M.O. Pintura", total_mo_p,  SUB1_FILL, W_FONT_SM)

            # Materiales por Superficie
            summary_row("Materiales por Superficie", mat_sup,   SUB2_FILL, W_FONT_IT, is_first=True)
            # Constante Material
            summary_row("Constante Material",         constante, SUB2_FILL, W_FONT_IT)
            # Total Materiales
            summary_row("Total Materiales",            total_mat, SUB2_FILL, W_FONT_SM)

            # Total M.O. y Materiales
            summary_row(f"TOTAL M.O. Y MATERIALES  ·  N° Orden {n_ord}",
                        total_mo_mat, SUB3_FILL, W_FONT_SM, is_first=True)

            grand_pin_total += total_mo_mat

        row += 1  # spacer

    # ── Gran total ────────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 20
    ws.merge_cells(f"A{row}:D{row}")
    ct = ws.cell(row,1,"  GRAN TOTAL  (todos los expedientes)  ·  M.O. Hojal/Mecánica + M.O. Pintura + Materiales")
    cv = ws.cell(row,5, grand_mo_total + grand_pin_total)
    for c in (ct,cv):
        c.font=Font(bold=True,size=11,color="FFFFFF")
        c.fill=GRAND_FILL; c.alignment=LFT if c==ct else RGT
        c.border=Border(left=THICK,right=THICK,top=THICK,bottom=THICK)
    ct.alignment=LFT; cv.alignment=RGT
    cv.number_format=MONEY

    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# UI
# ══════════════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown("### ⚙️ Opciones")
    show_meta = st.checkbox("Mostrar metadatos del expediente", value=True)
    st.markdown("---")
    st.markdown("**Formato:** AUDATEX / GNP Seguros")
    st.markdown("**N° Orden:** Referencia Interna del PDF")
    st.markdown("**Excel:** Una sola hoja por N° Orden")

uploaded = st.file_uploader(
    "📂 Arrastra uno o varios PDFs AUDATEX",
    type=["pdf"], accept_multiple_files=True,
)

if not uploaded:
    st.info("ℹ️ Sube al menos un PDF AUDATEX para comenzar.")
    st.stop()

all_data=[]; filenames=[]
prog = st.progress(0, text="Procesando…")
for i,f in enumerate(uploaded):
    prog.progress((i+1)/len(uploaded), text=f"Procesando: {f.name}")
    all_data.append(extract_audatex(f.read()))
    filenames.append(f.name)
prog.empty()

total_mo  = sum(d["meta"].get("total_mo",  sum(x["precio"] for x in d["mo"]))  for d in all_data)
total_pin = sum(d["meta"].get("total_mo_pintura", sum(x["precio"] for x in d["pintura"])) for d in all_data)
n_mo  = sum(len(d["mo"])      for d in all_data)
n_pin = sum(len(d["pintura"]) for d in all_data)

cols = st.columns(4)
for col,label,val in [
    (cols[0],"📄 PDFs procesados",    len(uploaded)),
    (cols[1],"🔵 Partidas M.O.",      n_mo),
    (cols[2],"🔴 Partidas Pintura",   n_pin),
    (cols[3],"💰 Total M.O.+Pintura", f"${total_mo+total_pin:,.2f}"),
]:
    col.markdown(f'<div class="metric-box"><div class="metric-value">{val}</div>'
                 f'<div class="metric-label">{label}</div></div>', unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

for fname,data in zip(filenames,all_data):
    meta=data["meta"]; mo=data["mo"]; pin=data["pintura"]
    n_ord=meta.get("num_orden","–")

    with st.expander(f"📄  N° Orden {n_ord}  —  {fname}", expanded=True):
        if show_meta:
            mc1,mc2,mc3=st.columns(3)
            mc1.markdown(f"**N° Orden:** {n_ord}")
            mc1.markdown(f"**Expediente:** {meta.get('expediente','–')}")
            mc2.markdown(f"**Taller:** {meta.get('taller','–')}")
            mc2.markdown(f"**Vehículo:** {meta.get('fabricante','')} {meta.get('modelo','')}")
            mc3.markdown(f"**VIN:** {meta.get('vin','–')}")
            st.divider()

        c1,c2=st.columns(2)

        with c1:
            st.markdown("**🔵 Mano de Obra Hojal/Mecánica**")
            if mo:
                df=pd.DataFrame(mo)
                df.insert(0,"N° Orden",n_ord)
                df.insert(1,"Sección","Mano de Obra")
                df.columns=["N° Orden","Sección","NR/Pos.","Trabajo","Precio ($)"]
                df["Precio ($)"]=df["Precio ($)"].map("${:,.2f}".format)
                st.dataframe(df,use_container_width=True,hide_index=True)
                t=meta.get("total_mo",sum(x["precio"] for x in mo))
                st.success(f"**Total M.O.: ${t:,.2f}**")
            else:
                st.warning("No se encontraron partidas.")

        with c2:
            st.markdown("**🔴 Pintura de Carrocería**")
            if pin:
                df=pd.DataFrame(pin)
                df.insert(0,"N° Orden",n_ord)
                df.insert(1,"Sección","Pintura")
                df.columns=["N° Orden","Sección","NR/Pos.","Descripción","Precio ($)"]
                df["Precio ($)"]=df["Precio ($)"].map("${:,.2f}".format)
                st.dataframe(df,use_container_width=True,hide_index=True)

                t_mo_p  = meta.get("tiempo_mo_pintura",  sum(x["precio"] for x in pin))
                t_prep  = meta.get("tiempo_prep_pintura", 0.0)
                t_mo_tot= meta.get("total_mo_pintura",    t_mo_p+t_prep)
                t_mat   = meta.get("total_materiales",    0.0)

                st.info(
                    f"Tiempo M.O.: **${t_mo_p:,.2f}** | "
                    f"Tiempo Prep.: **${t_prep:,.2f}** | "
                    f"Total M.O. Pintura: **${t_mo_tot:,.2f}**"
                )
                st.info(
                    f"Mat. Superficie: **${meta.get('mat_por_superficie',0):,.2f}** | "
                    f"Constante: **${meta.get('constante_material',0):,.2f}** | "
                    f"Total Materiales: **${t_mat:,.2f}**"
                )
                st.error(f"**Total M.O. y Materiales: ${t_mo_tot+t_mat:,.2f}**")
            else:
                st.warning("No se encontraron partidas.")

st.markdown("---")
st.markdown("### 📥 Exportar")
cx,cy=st.columns(2)

with cx:
    xlsx=build_excel(all_data,filenames)
    st.download_button(
        "⬇️ Descargar Excel (.xlsx)", data=xlsx,
        file_name="partidas_carroceria.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
with cy:
    rows=[]
    for fname,data in zip(filenames,all_data):
        n_ord=data["meta"].get("num_orden","–")
        for e in data["mo"]:
            rows.append({"N° Orden":n_ord,"Sección":"Mano de Obra",
                         "NR/Pos.":e["nr"],"Descripción":e["descripcion"],"Precio":e["precio"]})
        for e in data["pintura"]:
            rows.append({"N° Orden":n_ord,"Sección":"Pintura",
                         "NR/Pos.":e["nr"],"Descripción":e["descripcion"],"Precio":e["precio"]})
        # resumen pintura
        meta=data["meta"]
        for label,key in [
            ("Tiempo M.O.","tiempo_mo_pintura"),
            ("Tiempo Preparación","tiempo_prep_pintura"),
            ("Total M.O. Pintura","total_mo_pintura"),
            ("Materiales por Superficie","mat_por_superficie"),
            ("Constante Material","constante_material"),
            ("Total Materiales","total_materiales"),
        ]:
            if key in meta:
                rows.append({"N° Orden":n_ord,"Sección":"Resumen Pintura",
                             "NR/Pos.":"","Descripción":label,"Precio":meta[key]})
    csv=pd.DataFrame(rows).to_csv(index=False).encode()
    st.download_button(
        "⬇️ Descargar CSV", data=csv,
        file_name="partidas_MO H&P.csv", mime="text/csv",
        use_container_width=True,
    )
